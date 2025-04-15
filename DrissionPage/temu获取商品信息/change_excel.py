'''
=========================================================    
       @File     : change_excel.py
       @IDE      : PyCharm
       @Author   : Jing3
       @Date     : 2024/12/11 15:48
       @Desc     : 
=========================================================   
'''
from openpyxl import load_workbook


# 加载现有的 Excel 文件

# attr_list=['Skin Type',
# 'Specific Use',
# 'Active Ingredients',
# 'Item Form',
# 'Skin Care Ingredient',
# 'Product Attributes',
# 'Effect',
# 'Main Ingredients',
#            'Item ID']
# 判断是否找到大表头
# def change(attr_list):
# 传入Drissionpage获取到的属性列表
def change_excel(attr_list,path,key_list):
    wb = load_workbook("商品资料模板.xlsx")
    ws = wb.active

    # 定义大表头名称
    big_header_name = "商品属性(全英)"
    sec_header_name = "销售信息"

    # 遍历第一行，查找大表头名称
    big_header_cell = None
    for cell in ws[1]:  # 假设大表头在第一行
        if cell.value == big_header_name:
            big_header_cell = cell
            break

    sec_header_cell = None
    for cell2 in ws[1]:  # 假设大表头在第一行
        if cell2.value == sec_header_name:
            sec_header_cell = cell2
            break
    if big_header_cell:
        print(f"大表头 '{big_header_name}' 位于: 行 {big_header_cell.row}, 列 {big_header_cell.column}")
        print(f"结束表头 '{sec_header_name}' 位于: 行 {sec_header_cell.row}, 列 {sec_header_cell.column}")
        # 删除子表头（假设子表头在大表头下方的一行）
        header_row = big_header_cell.row + 1  # 子表头在大表头下方的一行
        for col in range(big_header_cell.column, sec_header_cell.column):  # 遍历所有列
            cell = ws.cell(row=header_row, column=col)
            cell.value = None  # 删除子表头单元格内容
        #写入属性民
        for col in range(big_header_cell.column, big_header_cell.column+len(attr_list)):
            cell = ws.cell(row=header_row, column=col)
            cell.value = attr_list[col-6]
            cell2 = ws.cell(row=header_row+1, column=col)
            cell2.value = key_list[col - 6]
            print(key_list[col-6])# 更改单元格内容
        # # 删除多余列
        # startcol=big_header_cell.column+len(attr_list)
        # duoyucol=sec_header_cell.column-(big_header_cell.column+len(attr_list))
        # ws.delete_cols(15, 11)
        # 保存修改后的 Excel 文件
        wb.save(f"{path}/商品资料模板.xlsx")
        wb.close()
        print("已删除 '商品属性' 下面的子表头内容。")
    else:
        print(f"没有找到大表头 '{big_header_name}'")
# ws.unmerge_cells('F1:Y1')
# ws.merge_cells('F1:N1')


