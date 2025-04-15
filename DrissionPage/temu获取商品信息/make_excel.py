import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

# 子表头
headers = [
    "Skin Type",
    "Specific Use",
    "Active Ingredients",
    "Item Form",
    "Skin Care Ingredient",
    "Product Attributes",
    "Effect",
    "Main Ingredients",
    "Item ID"
]

# 每列对应的初始宽度（可调整）
column_widths = [15, 20, 25, 12, 22, 20, 10, 25, 10]

# 创建一个空 DataFrame，并设置表头
df = pd.DataFrame(columns=headers)

# 创建一个 Excel 文件
wb = Workbook()
ws = wb.active
ws.title = "商品信息表"

# 蓝色填充
blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

# 边框（中等宽度）
medium_border = Border(
    left=Side(style="medium", color="000000"),
    right=Side(style="medium", color="000000"),
    top=Side(style="medium", color="000000"),
    bottom=Side(style="medium", color="000000")
)

# 表头字体加粗
bold_font = Font(bold=True)

# 写入大表头并设置背景颜色和边框
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
cell = ws.cell(row=1, column=1, value="商品信息(全英)")
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.fill = blue_fill
cell.border = medium_border
cell.font = bold_font

# 写入子表头并设置背景颜色、边框和加粗字体
for i, (header, width) in enumerate(zip(headers, column_widths), start=1):
    cell = ws.cell(row=2, column=i, value=header)
    cell.fill = blue_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = medium_border
    cell.font = bold_font
    # 设置列宽
    col_letter = ws.cell(row=2, column=i).column_letter
    ws.column_dimensions[col_letter].width = width

# 保存文件
output_path = "./商品信息表_蓝色背景_细边框.xlsx"
wb.save(output_path)

