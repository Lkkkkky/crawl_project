import os
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage  # 用于图片格式转换
# 读取 Excel 文件
file_path = 'Dropwiz主推商品.xlsx'  # 请根据实际路径修改
image_folder = "downloaded_images"  # 图片下载的文件夹
wb = load_workbook(file_path)
sheet = wb.active  # 假设使用的是第一个工作表，若有多个工作表，可使用 sheetnames 获取

# 获取第16列的图片 URL（假设数据从第二行开始，第一行为表头）
# sheet.max_row + 1
for row in range(2, 264):
    image_url = sheet.cell(row=row, column=3).value  # 获取第16列的图片URL

    if image_url and isinstance(image_url, str) and image_url.startswith(("http://", "https://")):
        try:
            # 下载图片
            response = requests.get(image_url)
            response.raise_for_status()  # 如果响应错误，抛出异常

            # 将图片加载到内存
            if response.status_code == 200:
                image_path = os.path.join(image_folder, f"image_{row}.webp")
                with open(image_path, "wb") as img_file:
                    img_file.write(response.content)

                # 将 .webp 图片转换为 .png
                png_image_path = os.path.join(image_folder, f"image_{row}.png")
                with PILImage.open(image_path) as img:
                    img.save(png_image_path, "PNG")
                # 创建 Image 对象
                img = Image(png_image_path)

                # 设置图片的宽度和高度为 100x80 像素
                img.width = 100
                img.height = 80

                # 获取单元格的当前列宽和行高（以便调整）
                column_letter = 'C'  # 假设第16列是 P 列
                row_height = sheet.row_dimensions[row].height

                # 调整列宽和行高以适应图片
                sheet.column_dimensions[column_letter].width = img.width / 7 -2 # 7是一个调整比例
                sheet.row_dimensions[row].height = img.height-20   # 微调行高，使图片不重叠

                # 设置图片锚定单元格的位置
                img.anchor = f'C{row}'  # 将图片插入到 P 列的当前行

                # 将图片插入到指定单元格
                sheet.add_image(img)

                # 清除原有的 URL 数据
                sheet.cell(row=row, column=3).value = None  # 清空原有的 URL 内容
                print(f'第{row}行插入成功')
        except requests.exceptions.RequestException as e:
            print(f"下载图片失败: {e}")
    else:
        print(f"无法下载图片，URL: {image_url}")
# 保存更新后的 Excel 文件
output_file_path = 'updated_阿里国际-商品-20250208_with_images_in_cell_resized.xlsx'
wb.save(output_file_path)


