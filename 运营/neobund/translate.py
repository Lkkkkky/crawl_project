'''
=========================================================    
       @File     : translate.py
       @IDE      : PyCharm
       @Author   : Jing3
       @Date     : 2024/12/11 16:14
       @Desc     : 
=========================================================   
'''
from openpyxl import load_workbook
from googletrans import Translator

# 加载 Excel 文件
wb = load_workbook("new_neobund分析.xlsx")
ws = wb.active

# 创建翻译器实例
translator = Translator()

# 假设我们要翻译整个工作表中的文本
for row in ws.iter_rows(min_col=1, max_col=1):  # 只遍历第二列
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            translated = translator.translate(cell.value.replace('[\'','').replace('\'','').replace('\']',''), src='zh-cn', dest='en')
            cell.value = translated.text

# 保存修改后的 Excel 文件
wb.save("translated_output.xlsx")
